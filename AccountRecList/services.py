from openpyxl import Workbook, load_workbook

#This function is called when a file is uploaded from a user.
#The function searches within the excel workbook on the summary tab 
#for a number to the left of the endingBalanceIdentifier tag
def Get_Reconciled_Balance(month, year, entity, accountNumber, pathname):
    endingBalanceIdentifier = "/EB"  + str(month) + str(year) + str(entity) + str(accountNumber)
    wb = load_workbook(filename = pathname, read_only=False, data_only=True)
    sheet_ranges = wb['Summary']

    for row in sheet_ranges.iter_rows():
        for entry in row:
            try:
                if endingBalanceIdentifier == entry.value:
                    endingBalanceIdentifier_Cell = entry
                    return endingBalanceIdentifier_Cell.offset(row=0, column=-1).value
                else:
                    return 0
            except(AttributeError, TypeError):
                continue

